# python -m venv env
# .\env\Scripts\activate


import pandas as pd
import subprocess
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
import os

# Path to your Excel file
excel_file_path = 'SampleData.xlsx'
excel_new_sheet_name = 'Max Forces By Part'
excluded_sheets = ["Parameters", "Results", "Statistics", excel_new_sheet_name]

num_before_stop = 54

# Open the Excel workbook
with pd.ExcelFile(excel_file_path) as xls:
    # Dictionary to hold max forces by part number
    max_forces = []
    
    # Iterate through the sheets
    for sheet_name in xls.sheet_names:  
        if sheet_name not in excluded_sheets:
            # Read the sheet into a DataFrame
            df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=3)

            # Find the index of the first row with force greater than 7.5
            hard_stop_index = df[df.iloc[:, 1] > 7.5].index.min()
            
            # If a hard stop is found, we select the data before it; otherwise, use all data
            if pd.notnull(hard_stop_index):
                # Select the data before the hard stop
                df_before_hard_stop = df.loc[:(hard_stop_index - num_before_stop)]
            else:
                df_before_hard_stop = df
            
            # Find the max force in the second column before the hard stop
            max_force_before_hard_stop = df_before_hard_stop.iloc[:, 1].max()
            
            # Assume part number is the prefix of the sheet name before the '.'
            part_number, run_number = sheet_name.split('.')   

            max_forces.append((part_number, run_number, max_force_before_hard_stop))

# Convert the dictionary to a DataFrame
max_forces_df = pd.DataFrame(max_forces, columns=['Part Number', 'Run Number', 'Max Force'])

# Load the workbook and add a new sheet
wb = load_workbook(excel_file_path)

# Check if the 'Max Forces By Part' sheet already exists
if excel_new_sheet_name in wb.sheetnames:
    # Use the existing sheet
    ws = wb[excel_new_sheet_name]
    # Clear the sheet
    ws.delete_rows(1, ws.max_row)

else:
    # Create a new sheet and place it at the first position
    ws = wb.create_sheet(title=excel_new_sheet_name, index=0)

# Write the DataFrame to the new sheet
for r_idx, row in enumerate(dataframe_to_rows(max_forces_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook
wb.save(excel_file_path)

print("Max forces by part number have been added to the Excel file.")


def get_full_file_path(file_name):
    # Get the current directory of the script
    current_dir = os.getcwd()
    # Combine the current directory with the file name
    full_path = os.path.join(current_dir, file_name)
    # Get the absolute path of the file
    absolute_path = os.path.abspath(full_path)
    return absolute_path

# Path to your JMP script (assuming it's ready to read the Excel file and knows which sheet to look for)
jmp_script_path = 'create_graph_template.jsl'


# Variables to pass to the JMP script
col1_name = 'Part Number'
col2_name = 'Max Force'
excel_file = 'SampleData.xlsx'
excel_file_path = get_full_file_path(excel_file)

image_name = "my_graph_2.png"
image_file_path = get_full_file_path(image_name)
print(image_file_path)

# Read in the JMP script template
with open(jmp_script_path, 'r') as file:
    jmp_script = file.read()

# Replace placeholders with actual values
jmp_script = jmp_script.replace('COL1NAME_PLACEHOLDER', col1_name)
jmp_script = jmp_script.replace('COL2NAME_PLACEHOLDER', col2_name)
jmp_script = jmp_script.replace('EXCEL_FILE_PATH_PLACEHOLDER', excel_file_path)
jmp_script = jmp_script.replace('SHEETNAME_PLACEHOLDER', excel_new_sheet_name)
jmp_script = jmp_script.replace('IMAGE_FILE_PATH_PLACEHOLDER', image_file_path)


# Write the modified script to a new file
modified_script_path = 'create_graph.jsl'
with open(modified_script_path, 'w') as file:
    file.write(jmp_script)

# Path to JMP executable
jmp_executable_path = 'C:/Program Files/SAS/JMP/16/jmp.exe'

# Command to run the modified JMP script
jmp_command = [jmp_executable_path, modified_script_path]

# Execute the command
subprocess.run(jmp_command, check=True)
print('Process completed successfully.')



